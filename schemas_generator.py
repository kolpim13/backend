import pandas as pd
import numpy as np
import openpyxl
from typing import List
from pathlib import Path

FOLDER_PATH = Path("C:\\projects\\impakt\\schemas_generator",)
EXCEL_PATH = Path(FOLDER_PATH, "shchemas.xlsx")

# Excel, Python, C#
TYPE_MAP = {
    'str': ('str', 'string'),
    'int': ('int', 'int'),
    'numeric': ('Decimal', 'decimal'),
    'bool': ('bool', 'bool'),
    'date': ('date', 'DateOnly'),
    'datetime': ('datetime', 'DateTime'),
    # Add more types as needed
}

DEFAULT_MAP = {
    'NULL': ('None', 'null'),
    'TRUE': ('True', 'true'),
    'FALSE': ('False', 'false'),
}

def read_named_table(ws, table_name: str) -> pd.DataFrame:
    for tbl in ws._tables.values():
            if tbl.name == table_name:
                data = ws[tbl.ref]
                headers = [cell.value for cell in data[0]]
                rows = [[cell.value for cell in row] for row in data[1:]]
                return pd.DataFrame(rows, columns=headers)

def DataFrame_to_fields(data_frame: pd.DataFrame) -> dict:
    VAR_NAME_POS = 0
    VAR_TYPE_POS = 1
    VAR_OPT_OPS = 2
    VAR_DEFAULT_POS = 3

    fields = []
    for frame in data_frame.values:
        field = {}
        field["name"] = frame[VAR_NAME_POS]
        field["type"] = frame[VAR_TYPE_POS]
        field["opt"] = frame[VAR_OPT_OPS]
        field["default"] = frame[VAR_DEFAULT_POS]
        fields.append(field)
    return fields

def generate_python_class(name: str, fields: dict) -> str:
    lines = ["class {name}(BaseModel):\n".format(name=name)]
    for field in fields:
        name = field["name"]
        type = TYPE_MAP[field["type"]][0]
        opt_open = "Optional[" if field["opt"] is True else ""
        opt_close = "]" if field["opt"] is True else ""
        value = field["default"] if field["default"] else ""
        value = " = " + "{}".format(DEFAULT_MAP[value][0]) if value in DEFAULT_MAP else value

        line = "    {name}: {opt_open}{type}{opt_close}{value}\n".format(name=name,
                                                                       opt_open=opt_open,
                                                                       type=type,
                                                                       opt_close=opt_close,
                                                                       value=value)
        lines.append(line)

    return lines

def generate_csharp_class(name: str, fields: dict) -> str:
    lines = ["public class {name}\n".format(name=name), "{\n"]
    for field in fields:
        name = field["name"]
        type = TYPE_MAP[field["type"]][1]
        required = "required " if field["opt"] is False else ""
        # required = ""
        opt = "?" if field["opt"] else ""
        value = field["default"] if field["default"] else ""
        value = " = {};".format(DEFAULT_MAP[value][1]) if value in DEFAULT_MAP else value
        line = "    {required}public {type}{opt} {name} {{ get; set; }}{value}\n".format(required=required,
                                                                                          type=type,
                                                                                          name=name,
                                                                                          opt=opt,
                                                                                          value=value)
        lines.append(line)

    lines.append("}")
    return lines

if __name__ == '__main__':
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    t = wb.worksheets
    ws = wb.active

    # Go through all worksheets - every worksheet <==> separate file.
    for ws in wb.worksheets:
        # Ommit template sheet
        if ws.title == "template":
            continue

        # Create new files [With the name of worksheet].
        path_py = Path(FOLDER_PATH, ws.title + ".py")
        path_cs = Path(FOLDER_PATH, ws.title + ".cs") 
        with open(path_py, 'w') as file_py, open(path_cs, 'w') as file_cs: 
            # Go through every table in sheet
            for table in ws._tables.values():
                # Get all data from a table
                data = ws[table.ref]
                headers = [cell.value for cell in data[0]]
                rows = [[cell.value for cell in row] for row in data[1:]]
                frame = pd.DataFrame(rows, columns=headers)

                # Convert it to apropriae format
                fields: dict = DataFrame_to_fields(frame)

                # Convert to strings
                class_name = table.name
                schema_python = generate_python_class(class_name, fields)
                schema_csharp = generate_csharp_class(class_name, fields)

                # Write data into files
                file_py.writelines(schema_python)
                file_py.write("\n")
                
                file_cs.writelines(schema_csharp)
                file_cs.write("\n")
